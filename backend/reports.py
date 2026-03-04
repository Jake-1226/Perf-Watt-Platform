"""
Report generation — Excel workbook with multiple sheets + summary statistics.
"""

import json
import sqlite3
import statistics
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from . import telemetry


HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(bold=True, size=14, color="2F5496")
DATA_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def _style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def generate_excel_report(db_path: str, output_path: str, run_metadata: dict = None):
    """Generate a comprehensive Excel report from telemetry data."""
    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary_sheet(ws_summary, db_path, run_metadata or {})

    # ── Sheet 2: OS Metrics (raw) ──
    ws_os = wb.create_sheet("OS Metrics")
    _write_os_metrics_sheet(ws_os, db_path)

    # ── Sheet 3: Power Metrics (raw) ──
    ws_power = wb.create_sheet("Power Metrics")
    _write_power_metrics_sheet(ws_power, db_path)

    # ── Sheet 4: Phase Summary ──
    ws_phases = wb.create_sheet("Phase Summary")
    _write_phase_summary_sheet(ws_phases, db_path)

    # ── Sheet 5: System Info ──
    ws_sysinfo = wb.create_sheet("System Info")
    _write_sysinfo_sheet(ws_sysinfo, db_path)

    # ── Sheet 6: Benchmark Events ──
    ws_events = wb.create_sheet("Benchmark Events")
    _write_events_sheet(ws_events, db_path)

    # ── Sheet 7: Charts ──
    ws_charts = wb.create_sheet("Charts")
    _write_charts_sheet(ws_charts, ws_os, ws_power, db_path)

    wb.save(output_path)
    return output_path


def _write_summary_sheet(ws, db_path, metadata):
    ws.cell(row=1, column=1, value="Performance Test Report").font = Font(bold=True, size=18, color="2F5496")
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = DATA_FONT
    ws.cell(row=3, column=1, value=f"Run ID: {metadata.get('run_id', 'N/A')}").font = DATA_FONT

    row = 5
    ws.cell(row=row, column=1, value="System Configuration").font = SECTION_FONT
    row += 1

    conn = sqlite3.connect(db_path)
    c = conn.cursor()

    # System info
    c.execute("SELECT source, key, value FROM system_info ORDER BY source, id")
    for r in c.fetchall():
        ws.cell(row=row, column=1, value=f"{r[0]}: {r[1]}").font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value=str(r[2])[:100]).font = DATA_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Test Overview").font = SECTION_FONT
    row += 1

    # Phase stats
    c.execute("""SELECT phase, COUNT(*) as samples,
                 AVG(cpu_pct) as avg_cpu, MAX(cpu_pct) as max_cpu,
                 AVG(mem_pct) as avg_mem
                 FROM os_metrics WHERE phase != '' GROUP BY phase ORDER BY MIN(epoch)""")
    os_phases = c.fetchall()

    c.execute("""SELECT phase, COUNT(*) as samples,
                 AVG(sys_input_ac_w) as avg_ac, MAX(sys_input_ac_w) as max_ac,
                 AVG(cpu_power_w) as avg_cpu_w
                 FROM power_metrics WHERE phase != '' GROUP BY phase ORDER BY MIN(epoch)""")
    pwr_phases = c.fetchall()

    headers = ["Phase", "OS Samples", "Avg CPU%", "Max CPU%", "Avg Mem%",
               "Pwr Samples", "Avg AC(W)", "Max AC(W)", "Avg CPU Pwr(W)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header(ws, row, len(headers))
    row += 1

    pwr_dict = {r[0]: r for r in pwr_phases}
    for osr in os_phases:
        phase = osr[0]
        pr = pwr_dict.get(phase, (phase, 0, None, None, None))
        vals = [phase, osr[1],
                round(osr[2], 1) if osr[2] else "N/A",
                round(osr[3], 1) if osr[3] else "N/A",
                round(osr[4], 1) if osr[4] else "N/A",
                pr[1],
                round(pr[2], 1) if pr[2] else "N/A",
                round(pr[3], 1) if pr[3] else "N/A",
                round(pr[4], 1) if pr[4] else "N/A"]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
        row += 1

    # Overall stats
    row += 1
    ws.cell(row=row, column=1, value="Overall Statistics").font = SECTION_FONT
    row += 1

    c.execute("SELECT MIN(epoch), MAX(epoch), COUNT(*) FROM os_metrics")
    os_stats = c.fetchone()
    if os_stats[0] and os_stats[1]:
        duration = os_stats[1] - os_stats[0]
        ws.cell(row=row, column=1, value="Total Duration").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{duration:.0f}s ({duration/60:.1f}m)")
        row += 1
        ws.cell(row=row, column=1, value="OS Metric Samples").font = Font(bold=True)
        ws.cell(row=row, column=2, value=os_stats[2])
        row += 1

    c.execute("SELECT COUNT(*) FROM power_metrics")
    ws.cell(row=row, column=1, value="Power Metric Samples").font = Font(bold=True)
    ws.cell(row=row, column=2, value=c.fetchone()[0])
    row += 1

    c.execute("SELECT AVG(sys_input_ac_w), MIN(sys_input_ac_w), MAX(sys_input_ac_w) FROM power_metrics WHERE sys_input_ac_w IS NOT NULL")
    pwr_overall = c.fetchone()
    if pwr_overall[0]:
        ws.cell(row=row, column=1, value="Avg AC Power").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{pwr_overall[0]:.1f}W")
        row += 1
        ws.cell(row=row, column=1, value="Min AC Power").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{pwr_overall[1]:.1f}W")
        row += 1
        ws.cell(row=row, column=1, value="Max AC Power").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{pwr_overall[2]:.1f}W")
        row += 1

    conn.close()
    _auto_width(ws)


def _write_os_metrics_sheet(ws, db_path):
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("SELECT * FROM os_metrics ORDER BY epoch ASC")
    cols = [d[0] for d in c.description]

    for i, col in enumerate(cols, 1):
        ws.cell(row=1, column=i, value=col)
    _style_header(ws, 1, len(cols))

    for row_idx, row_data in enumerate(c.fetchall(), 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
    conn.close()
    _auto_width(ws)


def _write_power_metrics_sheet(ws, db_path):
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("SELECT * FROM power_metrics ORDER BY epoch ASC")
    cols = [d[0] for d in c.description]

    for i, col in enumerate(cols, 1):
        ws.cell(row=1, column=i, value=col)
    _style_header(ws, 1, len(cols))

    for row_idx, row_data in enumerate(c.fetchall(), 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
    conn.close()
    _auto_width(ws)


def _write_phase_summary_sheet(ws, db_path):
    conn = sqlite3.connect(db_path)
    c = conn.cursor()

    ws.cell(row=1, column=1, value="Phase-by-Phase Power & Utilization Summary").font = SECTION_FONT

    headers = ["Phase", "Duration(s)", "Samples",
               "Avg CPU%", "Min CPU%", "Max CPU%",
               "Avg Mem%",
               "Avg AC(W)", "Min AC(W)", "Max AC(W)",
               "Avg CPU Pwr(W)", "Avg DIMM(W)", "Avg Storage(W)", "Avg Fan(W)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header(ws, 3, len(headers))

    # OS metrics by phase
    c.execute("""SELECT phase,
        MAX(epoch)-MIN(epoch) as dur, COUNT(*),
        AVG(cpu_pct), MIN(cpu_pct), MAX(cpu_pct), AVG(mem_pct)
        FROM os_metrics WHERE phase != '' GROUP BY phase ORDER BY MIN(epoch)""")
    os_data = {r[0]: r for r in c.fetchall()}

    # Power by phase
    c.execute("""SELECT phase, COUNT(*),
        AVG(sys_input_ac_w), MIN(sys_input_ac_w), MAX(sys_input_ac_w),
        AVG(cpu_power_w), AVG(dimm_power_w), AVG(storage_power_w), AVG(fan_power_w)
        FROM power_metrics WHERE phase != '' GROUP BY phase ORDER BY MIN(epoch)""")
    pwr_data = {r[0]: r for r in c.fetchall()}

    row = 4
    all_phases = list(dict.fromkeys(list(os_data.keys()) + list(pwr_data.keys())))
    for phase in all_phases:
        os = os_data.get(phase, (phase, 0, 0, None, None, None, None))
        pw = pwr_data.get(phase, (phase, 0, None, None, None, None, None, None, None))
        vals = [
            phase,
            round(os[1], 0) if os[1] else 0,
            os[2],
            round(os[3], 1) if os[3] else "N/A",
            round(os[4], 1) if os[4] else "N/A",
            round(os[5], 1) if os[5] else "N/A",
            round(os[6], 1) if os[6] else "N/A",
            round(pw[2], 1) if pw[2] else "N/A",
            round(pw[3], 1) if pw[3] else "N/A",
            round(pw[4], 1) if pw[4] else "N/A",
            round(pw[5], 1) if pw[5] else "N/A",
            round(pw[6], 1) if pw[6] else "N/A",
            round(pw[7], 1) if pw[7] else "N/A",
            round(pw[8], 1) if pw[8] else "N/A",
        ]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
        row += 1

    conn.close()
    _auto_width(ws)


def _write_sysinfo_sheet(ws, db_path):
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("SELECT collected_at, source, key, value FROM system_info ORDER BY source, id")

    headers = ["Collected At", "Source", "Key", "Value"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, 1, len(headers))

    for row_idx, r in enumerate(c.fetchall(), 2):
        for col_idx, val in enumerate(r, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(val)[:200]).font = DATA_FONT

    conn.close()
    _auto_width(ws)


def _write_events_sheet(ws, db_path):
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("SELECT timestamp, phase, event_type, benchmark, message FROM benchmark_events ORDER BY epoch ASC")

    headers = ["Timestamp", "Phase", "Event Type", "Benchmark", "Message"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, 1, len(headers))

    for row_idx, r in enumerate(c.fetchall(), 2):
        for col_idx, val in enumerate(r, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(val)[:200]).font = DATA_FONT

    conn.close()
    _auto_width(ws)


def _write_charts_sheet(ws_charts, ws_os, ws_power, db_path):
    """Add charts to the Charts sheet referencing OS and Power data sheets."""
    conn = sqlite3.connect(db_path)
    c = conn.cursor()

    # Count rows for chart ranges
    c.execute("SELECT COUNT(*) FROM os_metrics")
    os_count = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM power_metrics")
    pwr_count = c.fetchone()[0]
    conn.close()

    if os_count < 2:
        ws_charts.cell(row=1, column=1, value="Not enough data for charts").font = SECTION_FONT
        return

    # Chart 1: CPU Utilization over time
    chart1 = LineChart()
    chart1.title = "CPU Utilization Over Time"
    chart1.y_axis.title = "CPU %"
    chart1.x_axis.title = "Sample #"
    chart1.width = 30
    chart1.height = 15
    data_ref = Reference(ws_os, min_col=5, min_row=1, max_row=os_count + 1)
    chart1.add_data(data_ref, titles_from_data=True)
    ws_charts.add_chart(chart1, "A1")

    # Chart 2: Memory Utilization
    chart2 = LineChart()
    chart2.title = "Memory Utilization Over Time"
    chart2.y_axis.title = "Mem %"
    chart2.width = 30
    chart2.height = 15
    data_ref2 = Reference(ws_os, min_col=6, min_row=1, max_row=os_count + 1)
    chart2.add_data(data_ref2, titles_from_data=True)
    ws_charts.add_chart(chart2, "A18")

    if pwr_count >= 2:
        # Chart 3: AC Power
        chart3 = LineChart()
        chart3.title = "System AC Power Over Time"
        chart3.y_axis.title = "Watts"
        chart3.width = 30
        chart3.height = 15
        data_ref3 = Reference(ws_power, min_col=5, min_row=1, max_row=pwr_count + 1)
        chart3.add_data(data_ref3, titles_from_data=True)
        ws_charts.add_chart(chart3, "A35")

        # Chart 4: CPU Power
        chart4 = LineChart()
        chart4.title = "CPU Power Over Time"
        chart4.y_axis.title = "Watts"
        chart4.width = 30
        chart4.height = 15
        data_ref4 = Reference(ws_power, min_col=6, min_row=1, max_row=pwr_count + 1)
        chart4.add_data(data_ref4, titles_from_data=True)
        ws_charts.add_chart(chart4, "A52")


def generate_summary(db_path: str) -> dict:
    """Generate a JSON summary of the test run for the dashboard."""
    conn = sqlite3.connect(db_path)
    c = conn.cursor()

    summary = {"phases": [], "overall": {}, "system_info": {}}

    # Overall
    c.execute("SELECT MIN(epoch), MAX(epoch), COUNT(*) FROM os_metrics")
    r = c.fetchone()
    if r[0] and r[1]:
        summary["overall"]["duration_s"] = round(r[1] - r[0], 0)
        summary["overall"]["os_samples"] = r[2]

    c.execute("SELECT COUNT(*), AVG(sys_input_ac_w), MIN(sys_input_ac_w), MAX(sys_input_ac_w) FROM power_metrics WHERE sys_input_ac_w IS NOT NULL")
    r = c.fetchone()
    summary["overall"]["power_samples"] = r[0]
    if r[1]:
        summary["overall"]["avg_ac_w"] = round(r[1], 1)
        summary["overall"]["min_ac_w"] = round(r[2], 1)
        summary["overall"]["max_ac_w"] = round(r[3], 1)

    # Per-phase
    c.execute("""SELECT phase,
        COUNT(*), AVG(cpu_pct), MIN(cpu_pct), MAX(cpu_pct), AVG(mem_pct)
        FROM os_metrics WHERE phase != '' GROUP BY phase ORDER BY MIN(epoch)""")
    os_phases = {r[0]: r for r in c.fetchall()}

    c.execute("""SELECT phase,
        COUNT(*), AVG(sys_input_ac_w), MIN(sys_input_ac_w), MAX(sys_input_ac_w),
        AVG(cpu_power_w), AVG(dimm_power_w), AVG(storage_power_w), AVG(fan_power_w)
        FROM power_metrics WHERE phase != '' GROUP BY phase ORDER BY MIN(epoch)""")
    pwr_phases = {r[0]: r for r in c.fetchall()}

    all_phases = list(dict.fromkeys(list(os_phases.keys()) + list(pwr_phases.keys())))
    for phase in all_phases:
        os = os_phases.get(phase)
        pw = pwr_phases.get(phase)
        entry = {"name": phase}
        if os:
            entry["os_samples"] = os[1]
            entry["avg_cpu_pct"] = round(os[2], 1) if os[2] else None
            entry["max_cpu_pct"] = round(os[4], 1) if os[4] else None
            entry["avg_mem_pct"] = round(os[5], 1) if os[5] else None
        if pw:
            entry["pwr_samples"] = pw[1]
            entry["avg_ac_w"] = round(pw[2], 1) if pw[2] else None
            entry["avg_cpu_w"] = round(pw[5], 1) if pw[5] else None
            entry["avg_dimm_w"] = round(pw[6], 1) if pw[6] else None
            entry["avg_stor_w"] = round(pw[7], 1) if pw[7] else None
            entry["avg_fan_w"] = round(pw[8], 1) if pw[8] else None
        summary["phases"].append(entry)

    # System info
    c.execute("SELECT source, key, value FROM system_info")
    for r in c.fetchall():
        src = r[0]
        if src not in summary["system_info"]:
            summary["system_info"][src] = {}
        summary["system_info"][src][r[1]] = r[2]

    conn.close()
    return summary
